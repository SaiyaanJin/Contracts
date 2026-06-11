import os
import sys
import re
from datetime import datetime, date
import calendar
import openpyxl
from decimal import Decimal

# Add current directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from app import create_app
from app.extensions import db
from app.models.contract import Contract
from app.models.vendor import Vendor
from app.models.invoice import Invoice
from app.models.user import User

def extract_first_number(s):
    if s is None:
        return ""
    match = re.search(r'\d+[\d\.,]*', str(s))
    if match:
        return match.group(0)
    return ""

def clean_numeric_string(s):
    if s is None:
        return 0.0
    s = str(s).strip()
    if '=' in s:
        s = s.split('=')[-1]
    
    if '+' in s:
        parts = s.split('+')
        total = 0.0
        for p in parts:
            total += clean_numeric_string(p)
        return total
        
    num_str = extract_first_number(s)
    if not num_str:
        return 0.0
        
    match = re.search(r'[\.\,](\d{1,2})$', num_str)
    if match:
        decimal_part = match.group(1)
        integer_part = num_str[:-len(match.group(0))]
        integer_part = re.sub(r'[\.\,]', '', integer_part)
        try:
            return float(integer_part + '.' + decimal_part)
        except ValueError:
            return 0.0
    else:
        cleaned = re.sub(r'[\.\,]', '', num_str)
        try:
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0

def parse_date_safe(d_str):
    if not d_str:
        return None
    if isinstance(d_str, datetime):
        return d_str.date()
    if isinstance(d_str, date):
        return d_str
    
    d_str = str(d_str).strip()
    if d_str.upper() in ["PERPETUAL", "PERPETUAL CONTRACT", "FOREVER", "N/A", "NA", "-", ""]:
        if d_str.upper() in ["PERPETUAL", "PERPETUAL CONTRACT", "FOREVER"]:
            return date(2099, 12, 31)
        return None
    
    d_str = re.sub(r'[^\d\.\-\/]', '', d_str)
    
    formats = [
        "%d.%m.%Y", "%d-%m-%Y", "%d/%m/%Y",
        "%Y-%m-%d", "%d.%m.%y", "%d-%m-%y", "%d/%m/%y"
    ]
    for fmt in formats:
        try:
            return datetime.strptime(d_str, fmt).date()
        except ValueError:
            pass
            
    match = re.match(r'(\d{1,2})[\.\-\/](\d{1,2})[\.\-\/](\d{2,4})', d_str)
    if match:
        try:
            day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
            if year < 100:
                year += 2000
            month = max(1, min(12, month))
            _, max_days = calendar.monthrange(year, month)
            day = max(1, min(max_days, day))
            return date(year, month, day)
        except Exception:
            pass
        
    return None

def clean_vendor_name(name):
    if not name:
        return "Unknown Vendor"
    name = str(name).strip()
    name = re.sub(r'^(?i)M\/S\.?\s*', '', name)
    name = re.sub(r'^(?i)M\/s\.?\s*', '', name)
    name = re.sub(r'^(?i)M\/S\s*', '', name)
    name = re.sub(r'^(?i)M\/s\s*', '', name)
    return name.strip()

def get_or_create_vendor(name, vendor_map, admin_user):
    cleaned_name = clean_vendor_name(name)
    if cleaned_name in vendor_map:
        return vendor_map[cleaned_name]
    
    # Check database
    vendor = Vendor.query.filter_by(name=cleaned_name).first()
    if not vendor:
        # Generate new vendor code
        max_vendor = Vendor.query.order_by(Vendor.vendor_code.desc()).first()
        max_num = 1
        if max_vendor and max_vendor.vendor_code and max_vendor.vendor_code.startswith("VND-"):
            try:
                max_num = int(max_vendor.vendor_code.split("-")[1]) + 1
            except ValueError:
                pass
        
        vendor_code = f"VND-{max_num:05d}"
        vendor = Vendor(
            name=cleaned_name,
            vendor_code=vendor_code,
            status="Active",
            created_by_id=admin_user.id
        )
        db.session.add(vendor)
        db.session.flush() # get ID
        print(f"Created vendor: {cleaned_name} ({vendor_code})")
    
    vendor_map[cleaned_name] = vendor
    return vendor

def main():
    app = create_app()
    with app.app_context():
        print("Starting data ingestion...")
        
        # Get admin user
        admin_user = User.query.filter_by(emp_id="00001").first()
        if not admin_user:
            admin_user = User.query.filter_by(username="admin").first()
        if not admin_user:
            print("Admin user not found. Please run seed-db first.")
            return

        # Truncate existing data to prevent duplicate primary keys or weird behavior
        print("Clearing old invoices and contracts (keeping vendors, users, roles)...")
        db.session.query(Invoice).delete()
        db.session.query(Contract).delete()
        db.session.commit()
        
        # Load vendors map
        vendor_map = {}
        for v in Vendor.query.all():
            vendor_map[v.name] = v

        # ----------------------------------------------------
        # Part 1: Ingest Contracts
        # ----------------------------------------------------
        contract_files = [
            (r"E:\Contracts\List of All Contracts as on 31st may 21.xlsx", "Sheet1"),
            (r"E:\Contracts\list of contracts under freshor renewal process.xlsx", "Sheet1")
        ]
        
        contracts_by_no = {}
        
        for filepath, sheetname in contract_files:
            if not os.path.exists(filepath):
                print(f"Contract file not found: {filepath}")
                continue
                
            print(f"Processing contract register: {os.path.basename(filepath)}")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb[sheetname]
            rows = list(sheet.iter_rows(values_only=True))
            
            # Rows 6 onwards contain contracts (1-indexed row 6 is index 5)
            for r_idx in range(5, len(rows)):
                row = rows[r_idx]
                if not row or len(row) < 5 or not any(row):
                    continue
                
                # Check if this row is actually a contract row (PO/LoA details must be present)
                loa_no = row[1]
                if not loa_no:
                    continue
                
                loa_no = str(loa_no).strip()
                val_str = row[2]
                party_name = row[3] or "Unknown Vendor"
                from_date_raw = row[4]
                to_date_raw = row[5] if len(row) > 5 else None
                
                # Parse dates
                from_date = parse_date_safe(from_date_raw)
                to_date = parse_date_safe(to_date_raw)
                
                if not to_date:
                    to_date = date(2099, 12, 31) # Perpetual default
                if not from_date:
                    from_date = date(2020, 1, 1) # Fallback start
                
                vendor = get_or_create_vendor(party_name, vendor_map, admin_user)
                value = clean_numeric_string(val_str)
                
                # Try to extract loa date if present in text, e.g. "dated 06.11.2020"
                loa_date = None
                date_match = re.search(r'dated\s*(\d{1,2}[\.\-\/]\d{1,2}[\.\-\/]\d{2,4})', loa_no, re.IGNORECASE)
                if date_match:
                    loa_date = parse_date_safe(date_match.group(1))
                
                # Use LOA number as contract_no (clean it up slightly)
                contract_no = loa_no
                # If there's a newline, let's keep the first line as contract_no and full as name
                lines = contract_no.split('\n')
                c_no = lines[0].strip()
                name = c_no
                if len(lines) > 1:
                    name = lines[1].strip()
                else:
                    # If contract number is very long, make it short description
                    name = c_no[:200]
                
                # Avoid duplicates
                if c_no in contracts_by_no:
                    continue
                
                # Clean up name if it's too short or generic
                contract_name = f"IT Contract - {vendor.name}"
                if len(name) > 15:
                    contract_name = name
                
                contract = Contract(
                    contract_no=c_no,
                    file_no=f"ERLDC/IT/LEGACY/{r_idx+1:04d}",
                    name=contract_name,
                    department="Information Technology",
                    region="ERLDC HQ" if r_idx % 10 < 7 else ["ERLDC HQ", "Patna SLDC", "Bhubaneswar", "Ranchi", "Gangtok"][(r_idx // 10) % 4 + 1],
                    contract_type="Service" if "AMC" in contract_name or "ILL" in contract_name or "Leased" in contract_name else "Supply",
                    procurement_type="Offline",
                    estimated_value=Decimal(str(value)),
                    contract_value=Decimal(str(value)),
                    vendor_id=vendor.id,
                    start_date=from_date,
                    end_date=to_date,
                    loa_date=loa_date or from_date,
                    status="Active" if to_date >= date.today() else "Expired",
                    lifecycle_stage="Execution",
                    created_by_id=admin_user.id
                )
                db.session.add(contract)
                db.session.flush()
                contracts_by_no[c_no] = contract
                print(f"Added contract: {c_no} - {contract_name} ({value})")

        db.session.commit()
        print(f"Successfully loaded {len(contracts_by_no)} contracts from registers.")
        
        # ----------------------------------------------------
        # Part 2: Ingest Bills (Combined Registers)
        # ----------------------------------------------------
        bill_files = [
            r"E:\Bills\Bill Register 2025 Onwards.xlsx",
            r"E:\Bills\Old Bill Register till 2024.xlsx"
        ]
        
        bills_count = 0
        dyn_contracts_count = 0
        
        for filepath in bill_files:
            if not os.path.exists(filepath):
                print(f"Bill file not found: {filepath}")
                continue
                
            print(f"Processing bill register: {os.path.basename(filepath)}")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if 'Combine Bill Register' not in wb.sheetnames:
                print(f"Sheet 'Combine Bill Register' not found in {filepath}")
                continue
                
            sheet = wb['Combine Bill Register']
            rows = list(sheet.iter_rows(values_only=True))
            
            # Row 3 contains headers, Row 4 onwards contains bill data (1-indexed Row 4 is index 3)
            for r_idx in range(3, len(rows)):
                row = rows[r_idx]
                if not row or len(row) < 12 or not any(row):
                    continue
                
                bill_reg_no = row[0]
                if bill_reg_no is None:
                    continue
                    
                verification_date = parse_date_safe(row[1])
                eoffice_ref = row[2]
                party_name = row[3] or "Unknown Vendor"
                contract_name = row[4] or "IT Services"
                loa_ref = row[5]
                loa_date_raw = row[6]
                contract_type_lbl = row[7] # Regular / One time
                invoice_no = row[8]
                invoice_date = parse_date_safe(row[9])
                bill_period = row[10]
                billing_amount = clean_numeric_string(row[11])
                
                if not invoice_no:
                    invoice_no = f"REG-{bill_reg_no}"
                
                # Ensure we have a vendor
                vendor = get_or_create_vendor(party_name, vendor_map, admin_user)
                
                # Fallbacks for dates
                if not invoice_date:
                    invoice_date = verification_date or date.today()
                
                # Check if contract exists
                contract = None
                clean_loa = str(loa_ref).strip() if loa_ref else None
                
                if clean_loa:
                    contract = Contract.query.filter_by(contract_no=clean_loa).first()
                    if not contract:
                        # Also search by partial contract_no
                        contract = Contract.query.filter(Contract.contract_no.ilike(f"%{clean_loa}%")).first()
                
                if not contract and contract_name:
                    # Search by contract name and vendor
                    contract = Contract.query.filter(
                        Contract.name.ilike(f"%{contract_name}%"),
                        Contract.vendor_id == vendor.id
                    ).first()
                
                if not contract:
                    # Dynamically create contract
                    dyn_c_no = clean_loa if clean_loa else f"ERLDC/IT/GEN/{r_idx+1:04d}"
                    dyn_loa_date = parse_date_safe(loa_date_raw) or invoice_date
                    
                    contract = Contract(
                        contract_no=dyn_c_no,
                        file_no=str(eoffice_ref) if eoffice_ref else f"ERLDC/IT/DYN/{r_idx+1:04d}",
                        name=str(contract_name)[:500],
                        department="Information Technology",
                        region="ERLDC HQ" if r_idx % 10 < 7 else ["ERLDC HQ", "Patna SLDC", "Bhubaneswar", "Ranchi", "Gangtok"][(r_idx // 10) % 4 + 1],
                        contract_type="Service" if contract_type_lbl == "Regular" else "Supply",
                        procurement_type="Offline",
                        estimated_value=Decimal(str(billing_amount)),
                        contract_value=Decimal(str(billing_amount)),
                        vendor_id=vendor.id,
                        start_date=dyn_loa_date,
                        end_date=date(2099, 12, 31),
                        loa_date=dyn_loa_date,
                        status="Active",
                        lifecycle_stage="Execution",
                        created_by_id=admin_user.id
                    )
                    db.session.add(contract)
                    db.session.flush() # get ID
                    dyn_contracts_count += 1
                    print(f"Dynamically created contract: {dyn_c_no} for {vendor.name}")
                
                # Create invoice
                invoice = Invoice(
                    contract_id=contract.id,
                    vendor_id=vendor.id,
                    invoice_no=str(invoice_no)[:100],
                    vendor_invoice_no=str(invoice_no)[:100],
                    bill_type="Running Bill" if contract_type_lbl == "Regular" else "Final Bill",
                    invoice_amount=Decimal(str(billing_amount)),
                    gst_amount=Decimal("0.00"),
                    tds_amount=Decimal("0.00"),
                    net_payable=Decimal(str(billing_amount)),
                    invoice_date=invoice_date,
                    received_date=invoice_date,
                    verified_date=verification_date or invoice_date,
                    status="Paid",
                    created_by_id=admin_user.id
                )
                db.session.add(invoice)
                bills_count += 1
        
        db.session.commit()
        print(f"Data ingestion complete!")
        print(f"Total dynamic contracts created: {dyn_contracts_count}")
        print(f"Total bills/invoices populated: {bills_count}")

if __name__ == '__main__':
    main()
